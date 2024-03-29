# -*- coding: utf-8 -*-

from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import ReplyKeyboardRemove, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, ChatPermissions
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils.exceptions import Throttled
from aiogram.types.message import ContentType
import aiogram.utils.markdown as fmt
import asyncio

import soundfile as sf
import speech_recognition as sr

import logging
from time import time
import datetime
import random
from random import randint
import re
import json
import sqlite3
import requests
import numpy as np
import os
from config import *
import subprocess

import os.path

# import xlwt
from openpyxl import load_workbook, Workbook

#Database
conn = sqlite3.connect('db.db', check_same_thread=False)
cursor = conn.cursor()

def db(user_id: int, requests: int, data: str, admin: bool):
	cursor.execute('INSERT INTO main (user_id, requests, data, admin) VALUES (?, ?, ?, ?)', (user_id, requests, data, admin))
	conn.commit()
	
def req(user_id: int, data: str, text: str):
	cursor.execute('INSERT INTO requests (user_id, data, text) VALUES (?, ?, ?)', (user_id, data, text))
	conn.commit()

def get_data(): #Getting the current date
	now = datetime.datetime.now()
	data = str(now.year) + "-" + str(now.month) + "-" + str(now.day) + " " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second)
	return data

def audio_to_text(dest_name: str):
	r = sr.Recognizer()
	message = sr.AudioFile(dest_name)
	with message as source:
		audio = r.record(source)
	result = r.recognize_google(audio, language="ru_RU")
	return result

#initializing the bot
bot = Bot(token=token)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

#logg
logging.basicConfig(level=logging.INFO)

@dp.message_handler(commands = ["start"])
async def start(message: types.Message):
	cursor.execute(f"SELECT user_id FROM main WHERE user_id = '{message.from_user.id}'")
	if cursor.fetchone() is None:
		us_id = message.from_user.id
		us_name = message.from_user.first_name
		us_sname = message.from_user.last_name
		username = message.from_user.username
		db(user_id=us_id, requests=0, data=get_data(), admin=False)

	await message.answer(f"{message.from_user.first_name}, привет!\nЭтот бот умеет расшифровывать голосовые сообщения")

@dp.message_handler(commands =("help", "помощь"))
async def help(message: types.Message):
	buttons = [
		types.InlineKeyboardButton(text="Связь", url="https://t.me/Torkusz"),
		types.InlineKeyboardButton(text="Проект на GitHub", url="https://github.com/Torkusz/Decryption_of_voiceMessages")
	]
	keyboard = types.InlineKeyboardMarkup(row_width=1)
	keyboard.add(*buttons)
	await message.answer("Этот бот умеет расшифровывать голосовые сообщения\n\nДержи ссылку для обратной связи и на проект", reply_markup=keyboard)

@dp.message_handler(commands = ["get"])
async def start(message: types.Message):
	select_movies_query = f"SELECT user_id, data, text FROM requests WHERE user_id = '{message.from_user.id}'"
	cursor.execute(select_movies_query)
	result = cursor.fetchall()
	if result is None:
		await message.answer(f"{message.from_user.first_name}, кажется тебя нет в боте, нажми /start")
	else:
		await message.answer("📊Данные:\nЧтобы очистить нажмите /delete{}".format("".join(["\n"+str(f"DATA:%s  TEXT:%s"%(row[1],row[2])) for row in result])))

@dp.message_handler(commands="delete")
async def delite(message: types.Message):
	cursor.execute(f"DELETE FROM requests WHERE user_id = {message.from_user.id}")
	conn.commit()
	await message.answer("Удалил")

@dp.message_handler(commands=("me"))
async def me(message: types.Message):
	cursor.execute(f"SELECT * FROM main WHERE user_id={message.from_user.id}")
	result = cursor.fetchone()
	await message.answer(f"☃: @{message.from_user.username}\n📊Запросов - {result[1]}\n📅Дата регистрации - {result[2]}\nAdmin - {result[3]}", parse_mode="HTML", disable_web_page_preview=True)

@dp.message_handler(commands="file")
async def delite(message: types.Message):
	doc = open('xl/' + str(message.from_user.id) + '.xlsx', 'rb')
	await message.reply_document(doc)

@dp.message_handler(content_types=ContentType.TEXT)
async def check(message: types.Message):
	if message.text.startswith("/add"):
		try:
			cursor.execute(f"UPDATE main SET requests=requests+1 WHERE user_id = {message.from_user.id}")
			conn.commit()
			text = message.text.lower()

			if os.path.exists('xl/' + str(message.from_user.id) + '.xlsx') == True:
				fn = 'xl/' + str(message.from_user.id) + '.xlsx'
				wb = load_workbook(fn)
				if text.startswith('купить') == True:
					ws2 = wb['Список покупок']
					ws2.append([(str(get_data())), text])
				elif text.startswith('сделать') == True:
					ws3 = wb['Задачи']
					ws3.append([(str(get_data())), text])
				else:
					ws1 = wb['Заметки']
					ws1.append([(str(get_data())), text])
				wb.save(fn)
				wb.close
			else:
				wb = Workbook()
				ws1 = wb.create_sheet("Заметки", -2)
				ws2 = wb.create_sheet("Список покупок", -1)
				ws3 = wb.create_sheet("Задачи", 0)
				if text.startswith('купить') == True:
					ws2 = wb['Список покупок']
					ws2.append([(str(get_data())), text])
				elif text.startswith('сделать') == True:
					ws3 = wb['Задачи']
					ws3.append([(str(get_data())), text])
				else:
					ws1 = wb['Заметки']
					ws1.append([(str(get_data())), text])
				wb.save('xl/' + str(message.from_user.id) + '.xlsx')
				wb.close

			await message.answer(f"Записал:\n{text}")
		except sr.UnknownValueError as e:
			await message.answer("Прошу прощения, но я не разобрал сообщение, или оно поустое...")
	
@dp.message_handler(content_types=ContentType.VOICE)
async def check(message: types.Message):
	try:
		id = await bot.send_message(message.chat.id, f"Слушаю и понимаю...")
		file_info = await bot.get_file(message.voice.file_id)
		path = file_info.file_path
		fname = os.path.basename(path)
		
		doc = requests.get('https://api.telegram.org/file/bot{0}/{1}'.format(token, file_info.file_path))
		
		with open(fname+'.oga', 'wb') as f:
			f.write(doc.content)
		
		process = subprocess.run(['ffmpeg', '-i', fname+'.oga', fname+'.wav'])
		
		result = audio_to_text(fname+'.wav')
		# await bot.send_message(message.from_user.id, format(result))
		await bot.edit_message_text(chat_id=id.chat.id, message_id=id.message_id, text=f"Вот перевод:\n{result}")
		
		# os.remove(fname+'.wav')
		os.replace(f"{fname}.wav", f"audio/{fname}_{str(message.from_user.id)}.wav")
		os.remove(fname+'.oga')
		req(user_id=message.from_user.id, data=get_data(), text=result)
		cursor.execute(f"UPDATE main SET requests=requests+1 WHERE user_id = {message.from_user.id}")
		conn.commit()

		# book = xlwt.Workbook(encoding="utf-8")
		# sheet1 = book.add_sheet(str(message.from_user.id))
		# sheet1.write(1, 0, result)
		# book.save("test.xls")

		if os.path.exists('xl/' + str(message.from_user.id) + '.xlsx') == True:
			fn = 'xl/' + str(message.from_user.id) + '.xlsx'
			wb = load_workbook(fn)
			if result.startswith('купить') == True:
				ws2 = wb['Список покупок']
				ws2.append([(str(get_data())), result])
			elif result.startswith('сделать') == True:
				ws3 = wb['Задачи']
				ws3.append([(str(get_data())), result])
			else:
				ws1 = wb['Заметки']
				ws1.append([(str(get_data())), result])
			wb.save(fn)
			wb.close
		else:
			wb = Workbook()
			# ws = wb.active
			ws1 = wb.create_sheet("Заметки", -2)
			ws2 = wb.create_sheet("Список покупок", -1)
			ws3 = wb.create_sheet("Задачи", 0)
			if result.startswith('купить') == True:
				ws2 = wb['Список покупок']
				ws2.append([(str(get_data())), result])
			elif result.startswith('сделать') == True:
				ws3 = wb['Задачи']
				ws3.append([(str(get_data())), result])
			else:
				ws1 = wb['Заметки']
				ws1.append([(str(get_data())), result])
			wb.save('xl/' + str(message.from_user.id) + '.xlsx')
			wb.close

		# ws = wb['data']

	except sr.UnknownValueError as e:
		await message.answer("Прошу прощения, но я не разобрал сообщение, или оно поустое...")
	
if __name__ == "__main__":
	loop = asyncio.get_event_loop()
	executor.start_polling(dp, skip_updates=False)